from flask import Flask, render_template, request, jsonify, redirect, url_for, session
from extractor import data_manager
import sqlite3
import os
import csv
import re
from functools import wraps
from datetime import datetime, timedelta
import requests
import urllib3
import hashlib
try:
    import openpyxl
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

app = Flask(__name__)

# ========== CONFIGURACIÓN ==========
app.secret_key = os.environ.get('SECRET_KEY', 'dev_key_cambiar_en_produccion')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "database", "bolsa_datos.db")
DATA_DIR = os.path.join(BASE_DIR, "datos_dat")
FERIADOS_CSV = os.path.join(BASE_DIR, "feriados.csv")
DOLAR_XLSX = os.path.join(BASE_DIR, "dolar_bcv.xlsx")


def buscar_dat_reciente():
    if not os.path.exists(DATA_DIR):
        return None
    archivos = [(os.path.getmtime(os.path.join(DATA_DIR, f)), os.path.join(DATA_DIR, f))
                for f in os.listdir(DATA_DIR) if f.lower().endswith('.dat')]
    return max(archivos)[1] if archivos else None


def leer_ibc_de_dat(ruta_archivo):
    try:
        nombre = os.path.basename(ruta_archivo)
        match = re.search(r'(\d{8})', nombre)
        if match:
            f = match.group(1)
            fecha = f"{f[:4]}-{f[4:6]}-{f[6:]}" if int(f[:4]) > 1900 else f"{f[4:]}-{f[2:4]}-{f[:2]}"
        else:
            fecha = datetime.now().strftime('%Y-%m-%d')
        resultado = {'fecha': fecha, 'ibc': 0, 'ibc_var': 0, 'ibc_var_pct': 0, 'archivo': nombre}
        def conv(v):
            try:
                v = str(v).strip().replace('%','')
                if ',' in v and '.' in v: v = v.replace('.','').replace(',','.')
                elif ',' in v: v = v.replace(',','.')
                return float(v)
            except: return 0.0
        with open(ruta_archivo, 'r', encoding='latin-1') as fh:
            for linea in fh:
                if linea.startswith('IG|'):
                    partes = linea.strip().split('|')
                    if len(partes) >= 3: resultado['ibc'] = conv(partes[2])
                    if len(partes) >= 4: resultado['ibc_var'] = conv(partes[3])
                    if len(partes) >= 5: resultado['ibc_var_pct'] = conv(partes[4])
                    break
        return resultado
    except Exception as e:
        print(f"Error leyendo DAT: {e}")
        return None


def _fmt_fecha_xlsx(raw):
    f = str(int(raw))
    return f"{f[:4]}-{f[4:6]}-{f[6:]}" if len(f) == 8 else f


def leer_dolar_de_xlsx(fecha_str=None):
    if not OPENPYXL_DISPONIBLE or not os.path.exists(DOLAR_XLSX):
        return None
    try:
        wb  = openpyxl.load_workbook(DOLAR_XLSX, data_only=True)
        ws  = wb.active
        ultimo   = None
        objetivo = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]: continue
            fecha_fmt = _fmt_fecha_xlsx(row[0])
            tasa      = float(row[1])
            var_pct   = round(float(row[2]) * 100, 4) if row[2] else 0.0
            ultimo    = {'fecha': fecha_fmt, 'tasa': tasa, 'variacion_pct': var_pct}
            if fecha_str and fecha_fmt == fecha_str:
                objetivo = dict(ultimo)
        wb.close()
        return objetivo if (fecha_str and objetivo) else ultimo
    except Exception as e:
        print(f"Error leyendo Excel dolar: {e}")
        return None


def leer_todos_dolar_xlsx():
    if not OPENPYXL_DISPONIBLE or not os.path.exists(DOLAR_XLSX):
        return {}
    try:
        wb  = openpyxl.load_workbook(DOLAR_XLSX, data_only=True)
        ws  = wb.active
        resultado = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0] or not row[1]: continue
            resultado[_fmt_fecha_xlsx(row[0])] = float(row[1])
        wb.close()
        return resultado
    except Exception as e:
        print(f"Error leyendo Excel dolar todos: {e}")
        return {}


def normalizar_ibc(fechas, valores):
    """
    Normaliza la serie del IBC aplicando ÷1000 a los valores pre-rebase BVC.
    La BVC rebasó su índice el 2025-07-28 (de escala ~500.000 a ~500).
    Detecta el quiebre automáticamente (caída > 85% entre sesiones).
    """
    if len(valores) < 2:
        return valores
    idx_quiebre = None
    for i in range(1, len(valores)):
        prev = valores[i-1]
        curr = valores[i]
        if prev and curr and prev > 0:
            if (prev - curr) / prev > 0.85:
                idx_quiebre = i
                break
    if idx_quiebre is None:
        return valores
    print(f"⚙️  Rebase BVC detectado entre {fechas[idx_quiebre-1]} → {fechas[idx_quiebre]}. "
          f"Normalizando {idx_quiebre} registros ÷1000.")
    ajustados = list(valores)
    for j in range(idx_quiebre):
        if ajustados[j] is not None:
            ajustados[j] = round(ajustados[j] / 1000, 2)
    return ajustados

# Permisos disponibles para asignar a usuarios
PERMISOS_DISPONIBLES = {
    'ver_mercado':         'Ver Mercado Principal',
    'ver_rankings':        'Ver Rankings',
    'ver_analisis':        'Ver Análisis Técnico',
    'ver_consulta':        'Ver Consulta de Acciones',
    'ver_indices':         'Ver Índices',
    'ver_rankings_fechas': 'Ver Rankings por Fechas',
    'ver_portafolio':      'Ver Portafolio',
    'ver_prediccion':      'Ver Análisis Predictivo y Correlación',
}

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

# --- Feriados ---
def cargar_fechas_feriadas():
    fechas_feriadas = set()
    if os.path.exists(FERIADOS_CSV):
        try:
            with open(FERIADOS_CSV, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    try:
                        fecha_obj = datetime.strptime(row['Fecha'], '%d/%m/%Y')
                        fechas_feriadas.add(fecha_obj.strftime('%Y-%m-%d'))
                    except ValueError:
                        continue
        except Exception as e:
            print(f"Error cargando feriados: {e}")
    return fechas_feriadas

FECHAS_FERIADAS = cargar_fechas_feriadas()

# ========== HELPERS ==========
def format_spanish(value, decimals=2):
    try:
        if value is None: return "0,00"
        return f"{float(value):,.{decimals}f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    except: return str(value)

def tiene_imagen(simbolo):
    return os.path.exists(os.path.join(BASE_DIR, f"static/img/acciones/{simbolo}.png"))

app.jinja_env.globals.update(format_spanish=format_spanish, tiene_imagen=tiene_imagen)

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# ========== CREAR TABLA USUARIOS ==========
def crear_tabla_usuarios():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.execute('''CREATE TABLE IF NOT EXISTS usuarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        email TEXT,
        activo INTEGER DEFAULT 1,
        permisos TEXT DEFAULT '',
        fecha_registro TEXT DEFAULT CURRENT_TIMESTAMP
    )''')
    conn.commit()
    conn.close()

crear_tabla_usuarios()

# ========== CREDENCIALES ADMIN EN BD ==========
def _crear_tabla_admin():
    conn = sqlite3.connect(DB_PATH)
    conn.execute('''CREATE TABLE IF NOT EXISTS admin_credenciales (
        id INTEGER PRIMARY KEY CHECK (id=1),
        username TEXT NOT NULL DEFAULT 'admin',
        password TEXT NOT NULL
    )''')
    row = conn.execute("SELECT id FROM admin_credenciales WHERE id=1").fetchone()
    if not row:
        conn.execute(
            "INSERT INTO admin_credenciales (id, username, password) VALUES (1, 'admin', ?)",
            (hash_password('12345'),)
        )
        conn.commit()
    conn.close()

_crear_tabla_admin()

def get_admin_credenciales():
    conn = sqlite3.connect(DB_PATH)
    row = conn.execute("SELECT username, password FROM admin_credenciales WHERE id=1").fetchone()
    conn.close()
    return (row[0], row[1]) if row else ('admin', hash_password('12345'))

# ========== DECORADORES ==========
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_logged_in' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def permiso_requerido(permiso):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'admin_logged_in' in session:
                return f(*args, **kwargs)
            if 'user_logged_in' not in session:
                return redirect(url_for('login'))
            if permiso not in session.get('permisos', []):
                return render_template('sin_permiso.html'), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def api_permiso_requerido(permiso):
    """
    Igual que permiso_requerido pero para endpoints /api/...
    Devuelve JSON en lugar de redirect/HTML para no romper el fetch() del frontend.
    """
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'admin_logged_in' in session:
                return f(*args, **kwargs)
            if 'user_logged_in' not in session:
                return jsonify({"error": "no_session", "message": "Sesión expirada. Por favor inicia sesión nuevamente."}), 401
            if permiso not in session.get('permisos', []):
                return jsonify({"error": "sin_permiso", "message": "No tienes permiso para acceder a esta sección."}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

@app.context_processor
def inject_user():
    return dict(
        is_admin='admin_logged_in' in session,
        is_user='user_logged_in' in session,
        current_user=session.get('username'),
        user_permisos=session.get('permisos', [])
    )

# ========== NUEVO CONTEXT PROCESSOR PARA now() ==========
@app.context_processor
def inject_now():
    from datetime import datetime
    return {'now': datetime.now}

# ========== RSI ==========
def calcular_rsi(precios, periodo=14):
    if len(precios) < periodo: return 50.0
    subidas, bajadas = [], []
    for i in range(1, len(precios)):
        dif = precios[i] - precios[i-1]
        subidas.append(max(0, dif))
        bajadas.append(max(0, -dif))
    avg_g = sum(subidas[-periodo:]) / periodo
    avg_p = sum(bajadas[-periodo:]) / periodo
    if avg_p == 0: return 100.0
    return 100 - (100 / (1 + avg_g / avg_p))

def obtener_datos_bvc(simbolo):
    url = f"https://market.bolsadecaracas.com/api/mercado/resumen/simbolos/{simbolo}/libroOrdenes"
    try:
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, verify=False, timeout=5)
        if r.status_code == 200: return r.json()
    except: pass
    return None

# ========== AUTENTICACIÓN ==========

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        admin_user, admin_pass_hash = get_admin_credenciales()
        if username == admin_user and hash_password(password) == admin_pass_hash:
            session['admin_logged_in'] = True
            session['username'] = username
            return redirect(url_for('admin_panel'))

        conn = get_db()
        user = conn.execute(
            "SELECT * FROM usuarios WHERE username = ? AND activo = 1", (username,)
        ).fetchone()
        conn.close()

        if user and user['password'] == hash_password(password):
            session['user_logged_in'] = True
            session['username'] = username
            session['permisos'] = user['permisos'].split(',') if user['permisos'] else []
            return redirect(url_for('index'))

        return render_template('admin/login.html', error="Credenciales incorrectas")
    return render_template('admin/login.html')

@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        email    = request.form.get('email', '').strip()
        if not username or not password:
            return render_template('registro.html', error="Usuario y contraseña son obligatorios.")
        conn = get_db()
        try:
            conn.execute(
                "INSERT INTO usuarios (username, password, email, permisos) VALUES (?, ?, ?, ?)",
                (username, hash_password(password), email, '')
            )
            conn.commit()
            conn.close()
            return render_template('registro.html', success="¡Cuenta creada! Ya puedes iniciar sesión.")
        except sqlite3.IntegrityError:
            conn.close()
            return render_template('registro.html', error="Ese nombre de usuario ya está en uso.")
    return render_template('registro.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

# ========== ADMIN ==========

@app.route('/admin')
@login_required
def admin_panel():
    conn = get_db()
    usuarios = [dict(u) for u in conn.execute("SELECT * FROM usuarios ORDER BY fecha_registro DESC").fetchall()]
    conn.close()
    return render_template('admin/panel.html',
                           current_date=datetime.now().strftime('%Y-%m-%d'),
                           usuarios=usuarios,
                           permisos_disponibles=PERMISOS_DISPONIBLES)

@app.route('/admin/guardar', methods=['POST'])
@login_required
def admin_guardar():
    fecha    = request.form.get('fecha')
    simbolo  = request.form.get('simbolo').upper()
    nombre   = request.form.get('nombre')
    apertura = float(request.form.get('apertura', 0))
    cierre   = float(request.form.get('cierre', 0))
    maximo   = float(request.form.get('maximo', 0))
    minimo   = float(request.form.get('minimo', 0))
    volumen  = float(request.form.get('volumen', 0))
    variacion = ((cierre - apertura) / apertura) * 100 if apertura > 0 else 0
    conn = get_db()
    try:
        conn.execute("""INSERT OR REPLACE INTO mercado
            (fecha, simbolo, nombre, apertura, maximo, minimo, cierre, volumen, variacion)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (fecha, simbolo, nombre, apertura, maximo, minimo, cierre, volumen, variacion))
        conn.commit()
    finally:
        conn.close()
    return redirect(url_for('admin_panel'))

@app.route('/admin/usuarios/permisos', methods=['POST'])
@login_required
def admin_actualizar_permisos():
    data     = request.json
    user_id  = data.get('user_id')
    permisos = ','.join(data.get('permisos', []))
    conn = get_db()
    conn.execute("UPDATE usuarios SET permisos = ? WHERE id = ?", (permisos, user_id))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})

@app.route('/admin/usuarios/toggle', methods=['POST'])
@login_required
def admin_toggle_usuario():
    user_id = request.json.get('user_id')
    conn = get_db()
    conn.execute("UPDATE usuarios SET activo = CASE WHEN activo=1 THEN 0 ELSE 1 END WHERE id=?", (user_id,))
    conn.commit()
    row = conn.execute("SELECT activo FROM usuarios WHERE id=?", (user_id,)).fetchone()
    conn.close()
    return jsonify({"status": "ok", "activo": row['activo']})

@app.route('/admin/usuarios/eliminar', methods=['POST'])
@login_required
def admin_eliminar_usuario():
    conn = get_db()
    conn.execute("DELETE FROM usuarios WHERE id=?", (request.json.get('user_id'),))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})

# ========== RUTAS PÚBLICAS ==========

@app.route('/')
def index():
    conn = get_db()
    lista_fechas = [r[0] for r in conn.execute("SELECT DISTINCT fecha FROM mercado WHERE fecha >= '2024-01-01' ORDER BY fecha DESC").fetchall()]
    fecha_actual = request.args.get('fecha', lista_fechas[0] if lista_fechas else None)
    acciones, stats = [], {'suben': 0, 'bajan': 0, 'estables': 0, 'total': 0}
    if fecha_actual:
        rows = conn.execute("SELECT * FROM mercado WHERE fecha=? ORDER BY simbolo", (fecha_actual,)).fetchall()
        acciones = [dict(r) for r in rows]
        stats = {
            'total': len(acciones),
            'suben': sum(1 for a in acciones if a['variacion'] > 0),
            'bajan': sum(1 for a in acciones if a['variacion'] < 0),
            'estables': sum(1 for a in acciones if a['variacion'] == 0)
        }
    conn.close()
    return render_template('index.html', acciones=acciones, fecha_actual=fecha_actual,
                           lista_fechas=lista_fechas, estadisticas=stats)

@app.route('/analisis-tecnico')
@permiso_requerido('ver_analisis')
def vista_analisis_tecnico():
    return render_template('analisis_tecnico.html')

@app.route('/consulta')
@permiso_requerido('ver_consulta')
def consulta():
    conn = get_db()
    acciones = [dict(r) for r in conn.execute("""
        SELECT simbolo, nombre FROM mercado
        WHERE cierre > 0
        GROUP BY simbolo
        HAVING MAX(fecha) >= date('now', '-6 months')
        ORDER BY simbolo
    """).fetchall()]
    conn.close()
    return render_template('consulta.html', acciones=acciones)

@app.route('/rankings')
@permiso_requerido('ver_rankings')
def vista_rankings():
    periodos = {
        '1_Semana': {'label': '1 Semana',  'dias_objetivo': 7,   'margen': 3},
        '1_Mes':    {'label': '1 Mes',      'dias_objetivo': 30,  'margen': 5},
        '3_Meses':  {'label': '3 Meses',    'dias_objetivo': 90,  'margen': 8},
        '6_Meses':  {'label': '6 Meses',    'dias_objetivo': 180, 'margen': 10},
        '1_Año':    {'label': '1 Año',      'dias_objetivo': 365, 'margen': 15}
    }
    rankings_data = {}
    conn = get_db()
    ultima_fecha = conn.execute("SELECT MAX(fecha) FROM mercado").fetchone()[0]

    rows_dia = conn.execute(f"""
        SELECT simbolo, nombre, apertura, cierre, volumen,
               ((cierre - apertura) / apertura) * 100 as variacion
        FROM mercado WHERE fecha = '{ultima_fecha}' AND apertura > 0
    """).fetchall()
    data_dia = [dict(r) for r in rows_dia]

    rankings_data['Dia'] = {
        'titulo': 'Hoy', 'rango': f"Sesión del {ultima_fecha}",
        'ganadoras':  sorted([x for x in data_dia if x['variacion'] > 0.01], key=lambda x: x['variacion'], reverse=True)[:5],
        'perdedoras': sorted([x for x in data_dia if x['variacion'] < -0.01], key=lambda x: x['variacion'])[:5],
        'mas_negociadas':   sorted(data_dia, key=lambda x: x['volumen'] or 0, reverse=True)[:5],
        'menos_negociadas': sorted([x for x in data_dia if x['volumen'] > 0], key=lambda x: x['volumen'])[:5],
        'labels': [], 'valores': []
    }
    g = rankings_data['Dia']['ganadoras']; p = rankings_data['Dia']['perdedoras']
    rankings_data['Dia']['labels'] = [x['simbolo'] for x in g + p]
    rankings_data['Dia']['valores'] = [round(float(x['variacion']), 2) for x in g + p]

    for key, info in periodos.items():
        rows = conn.execute(f"""
            WITH RangoFechas AS (
                SELECT simbolo, SUM(volumen) as volumen_total FROM mercado
                WHERE julianday('{ultima_fecha}') - julianday(fecha) <= {info['dias_objetivo']} GROUP BY simbolo
            ),
            PrecioActual AS (SELECT simbolo, cierre FROM mercado WHERE fecha = '{ultima_fecha}'),
            PrecioPasado AS (
                SELECT simbolo, apertura, fecha,
                ROW_NUMBER() OVER (PARTITION BY simbolo ORDER BY ABS(julianday('{ultima_fecha}') - julianday(fecha) - {info['dias_objetivo']}) ASC) as rn
                FROM mercado
                WHERE julianday('{ultima_fecha}') - julianday(fecha)
                      BETWEEN {info['dias_objetivo'] - info['margen']} AND {info['dias_objetivo'] + info['margen']}
            )
            SELECT a.simbolo, a.cierre as actual, p.apertura as pasado, p.fecha as fecha_inicio,
                   ((a.cierre - p.apertura) / p.apertura) * 100 as rendimiento, rf.volumen_total
            FROM PrecioActual a
            JOIN PrecioPasado p ON a.simbolo = p.simbolo
            JOIN RangoFechas rf ON a.simbolo = rf.simbolo
            WHERE p.rn = 1 AND p.apertura > 0
        """).fetchall()
        data = [dict(r) for r in rows]
        g2 = sorted([x for x in data if x['rendimiento'] > 0.01], key=lambda x: x['rendimiento'], reverse=True)[:5]
        p2 = sorted([x for x in data if x['rendimiento'] < -0.01], key=lambda x: x['rendimiento'])[:5]
        rankings_data[key] = {
            'titulo': info['label'],
            'rango': f"Desde {data[0]['fecha_inicio'] if data else 'N/A'} hasta {ultima_fecha}",
            'ganadoras': g2, 'perdedoras': p2,
            'mas_negociadas':   sorted(data, key=lambda x: x['volumen_total'] or 0, reverse=True)[:5],
            'menos_negociadas': sorted([x for x in data if x['volumen_total'] > 0], key=lambda x: x['volumen_total'])[:5],
            'labels': [x['simbolo'] for x in g2 + p2],
            'valores': [round(float(x['rendimiento']), 2) for x in g2 + p2]
        }

    conn.close()
    return render_template('rankings.html', rankings=rankings_data)

# ========== RANKINGS POR FECHAS PERSONALIZADAS ==========

@app.route('/rankings-fechas')
@permiso_requerido('ver_rankings_fechas')
def vista_rankings_fechas():
    conn = get_db()
    fechas_disponibles = [r[0] for r in conn.execute(
        "SELECT DISTINCT fecha FROM mercado ORDER BY fecha DESC"
    ).fetchall()]
    conn.close()
    return render_template('rankings_fechas.html', fechas_disponibles=fechas_disponibles)

@app.route('/api/rankings-fechas')
def api_rankings_fechas():
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin    = request.args.get('fecha_fin')
    if not fecha_inicio or not fecha_fin:
        return jsonify({"error": "Parámetros incompletos"}), 400

    conn = get_db()
    rows = conn.execute("""
        WITH PrecioInicio AS (
            SELECT simbolo, nombre, apertura, fecha as f_inicio,
                   ROW_NUMBER() OVER (PARTITION BY simbolo ORDER BY fecha ASC) as rn
            FROM mercado WHERE fecha >= ? AND apertura > 0
        ),
        PrecioFin AS (
            SELECT simbolo, cierre, fecha as f_fin,
                   ROW_NUMBER() OVER (PARTITION BY simbolo ORDER BY fecha DESC) as rn
            FROM mercado WHERE fecha <= ?
        ),
        Volumen AS (
            SELECT simbolo, SUM(volumen) as volumen_total
            FROM mercado WHERE fecha BETWEEN ? AND ? GROUP BY simbolo
        )
        SELECT pi.simbolo, pi.nombre, pi.apertura as precio_inicio, pf.cierre as precio_fin,
               pi.f_inicio, pf.f_fin,
               ((pf.cierre - pi.apertura) / pi.apertura) * 100 as rendimiento,
               v.volumen_total
        FROM PrecioInicio pi
        JOIN PrecioFin pf ON pi.simbolo = pf.simbolo AND pi.rn=1 AND pf.rn=1
        JOIN Volumen v ON pi.simbolo = v.simbolo
        WHERE pi.apertura > 0
        ORDER BY rendimiento DESC
    """, (fecha_inicio, fecha_fin, fecha_inicio, fecha_fin)).fetchall()
    data = [dict(r) for r in rows]

    conn.close()
    return jsonify({
        "ganadoras":      sorted([x for x in data if x['rendimiento'] > 0.01], key=lambda x: x['rendimiento'], reverse=True)[:10],
        "perdedoras":     sorted([x for x in data if x['rendimiento'] < -0.01], key=lambda x: x['rendimiento'])[:10],
        "mas_negociadas": sorted(data, key=lambda x: x['volumen_total'] or 0, reverse=True)[:10],
        "menos_negociadas": sorted([x for x in data if x['volumen_total'] > 0], key=lambda x: x['volumen_total'])[:10],
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "total": len(data)
    })

# ========== ÍNDICES ==========

@app.route('/indices')
@permiso_requerido('ver_indices')
def vista_indices():
    return render_template('indices.html')

@app.route('/api/indices')
def api_indices():
    dias = request.args.get('dias', default=7, type=int)
    conn = get_db()
    indices = conn.execute('SELECT fecha, dolar, ibc FROM indicadores ORDER BY fecha DESC LIMIT ?', (dias,)).fetchall()
    conn.close()
    return jsonify([dict(ix) for ix in reversed(indices)])

@app.route('/api/ultimo_indice')
def api_ultimo_indice():
    conn = get_db()
    res = conn.execute('SELECT * FROM indicadores ORDER BY fecha DESC LIMIT 1').fetchone()
    conn.close()
    return jsonify(dict(res)) if res else jsonify({"ibc": 0, "dolar": 0, "fecha": "N/A"})

@app.route('/api/indices-hoy')
def api_indices_hoy():
    """Datos del día: IBC de tabla 'indices' (último registro) + dólar del Excel"""
    resultado = {
        'ibc': None, 'ibc_var': None, 'ibc_var_pct': None,
        'dolar': None, 'dolar_var_pct': None, 'ibc_usd': None,
        'fecha': None, 'fecha_dolar': None, 'archivo_dat': None,
        'fuente_ibc': None, 'fuente_dolar': None
    }

    def fecha_int_a_str(f):
        """Convierte 20251226 → '2025-12-26'"""
        f = str(int(f))
        return f"{f[:4]}-{f[4:6]}-{f[6:]}" if len(f) == 8 else f

    # ── IBC: último registro de tabla 'indicadores' ──
    conn = get_db()
    row_ibc = conn.execute(
        "SELECT fecha, ibc FROM indicadores WHERE ibc > 0 ORDER BY fecha DESC LIMIT 1"
    ).fetchone()
    row_ibc_prev = conn.execute(
        "SELECT ibc FROM indicadores WHERE ibc > 0 ORDER BY fecha DESC LIMIT 1 OFFSET 1"
    ).fetchone()
    conn.close()

    if row_ibc and row_ibc['ibc']:
        resultado['fecha']      = row_ibc['fecha']
        resultado['ibc']        = round(row_ibc['ibc'], 2)
        resultado['fuente_ibc'] = 'db'
        resultado['ibc_var']    = 0
        if row_ibc_prev and row_ibc_prev['ibc'] and row_ibc_prev['ibc'] > 0:
            resultado['ibc_var_pct'] = round(
                (row_ibc['ibc'] - row_ibc_prev['ibc']) / row_ibc_prev['ibc'] * 100, 2
            )
        else:
            resultado['ibc_var_pct'] = 0.0

    # ── Fallback: archivo .dat si no hay datos en BD ──
    if resultado['ibc'] is None:
        ruta_dat = buscar_dat_reciente()
        if ruta_dat:
            datos_dat = leer_ibc_de_dat(ruta_dat)
            if datos_dat and datos_dat['ibc'] > 0:
                resultado['ibc']         = round(datos_dat['ibc'], 2)
                resultado['ibc_var']     = round(datos_dat['ibc_var'], 2)
                resultado['ibc_var_pct'] = round(datos_dat['ibc_var_pct'], 2)
                resultado['fecha']       = datos_dat['fecha']
                resultado['archivo_dat'] = datos_dat['archivo']
                resultado['fuente_ibc']  = 'dat'

    # ── Dólar: primero tabla indicadores, luego fallback Excel ──
    conn2 = get_db()
    row_dolar = conn2.execute(
        "SELECT fecha, dolar FROM indicadores WHERE dolar > 0 ORDER BY fecha DESC LIMIT 1"
    ).fetchone()
    row_dolar_prev = conn2.execute(
        "SELECT dolar FROM indicadores WHERE dolar > 0 ORDER BY fecha DESC LIMIT 1 OFFSET 1"
    ).fetchone()
    conn2.close()

    if row_dolar and row_dolar['dolar']:
        dolar_val  = round(row_dolar['dolar'], 4)
        dolar_prev = row_dolar_prev['dolar'] if row_dolar_prev else None
        var_pct    = round((dolar_val - dolar_prev) / dolar_prev * 100, 4) if dolar_prev and dolar_prev > 0 else 0.0
        resultado['dolar']         = dolar_val
        resultado['dolar_var_pct'] = var_pct
        resultado['fecha_dolar']   = row_dolar['fecha']
        resultado['fuente_dolar']  = 'bd'
    else:
        fecha_buscar = resultado.get('fecha')
        datos_dolar  = leer_dolar_de_xlsx(fecha_buscar) if fecha_buscar else None
        if not datos_dolar:
            datos_dolar = leer_dolar_de_xlsx()
        if datos_dolar:
            resultado['dolar']         = round(datos_dolar['tasa'], 4)
            resultado['dolar_var_pct'] = round(datos_dolar.get('variacion_pct', 0), 4)
            resultado['fecha_dolar']   = datos_dolar['fecha']
            resultado['fuente_dolar']  = 'xlsx'

    # ── IBC en USD ──
    if resultado['ibc'] and resultado['dolar'] and resultado['dolar'] > 0:
        resultado['ibc_usd'] = round(resultado['ibc'] / resultado['dolar'], 4)

    return jsonify(resultado)


@app.route('/api/comparativa-indices')
def api_comparativa_indices():
    """
    Serie histórica de IBC desde tabla 'indices'
    + Dólar desde tabla 'indicadores' (con fallback al Excel).
    """
    def fecha_int_a_str(f):
        f = str(int(f))
        return f"{f[:4]}-{f[4:6]}-{f[6:]}" if len(f) == 8 else f

    conn = get_db()
    rows = conn.execute(
        "SELECT fecha, ibc as valor FROM indicadores WHERE ibc > 0 ORDER BY fecha ASC"
    ).fetchall()
    dolar_rows = conn.execute(
        "SELECT fecha, dolar FROM indicadores WHERE dolar > 0 ORDER BY fecha ASC"
    ).fetchall()
    conn.close()

    dolar_bd   = {r['fecha']: r['dolar'] for r in dolar_rows}
    dolar_xlsx = leer_todos_dolar_xlsx()

    fechas        = []
    valores_ibc   = []
    valores_dolar = []

    for r in rows:
        fecha_str = r['fecha']
        fechas.append(fecha_str)
        valores_ibc.append(round(r['valor'], 2))
        valores_dolar.append(dolar_bd.get(fecha_str) or dolar_xlsx.get(fecha_str) or None)

    valores_ibc = normalizar_ibc(fechas, valores_ibc)

    return jsonify({
        'fechas':        fechas,
        'valores_ibc':   valores_ibc,
        'valores_dolar': valores_dolar
    })
# ========== ADMIN: TABLA INDICES ==========

@app.route('/admin/guardar_indice_manual', methods=['POST'])
@login_required
def admin_guardar_indice_manual():
    """Guarda o actualiza un registro IBC en la tabla indices"""
    data      = request.json
    fecha_str = str(data.get('fecha', '')).replace('-', '')   # YYYY-MM-DD → YYYYMMDD int
    valor     = float(data.get('valor', 0))
    variacion = float(data.get('variacion', 0))
    fuente    = data.get('fuente', 'manual')
    if not fecha_str or not valor:
        return jsonify({"status": "error", "message": "Fecha y valor son requeridos"}), 400
    fecha_fmt = f"{fecha_str[:4]}-{fecha_str[4:6]}-{fecha_str[6:]}" if len(fecha_str)==8 else fecha_str
    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO indicadores (fecha, ibc, dolar)
            VALUES (?, ?, 0)
            ON CONFLICT(fecha) DO UPDATE SET ibc=excluded.ibc
        """, (fecha_fmt, valor))
        conn.commit()
        conn.close()
        return jsonify({"status": "ok", "fecha": fecha_fmt, "valor": valor})
    except Exception as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/admin/eliminar_indice_manual', methods=['POST'])
@login_required
def admin_eliminar_indice_manual():
    """Elimina un registro de la tabla indices por fecha"""
    data      = request.json
    fecha_str = str(data.get('fecha', '')).replace('-', '')
    fecha_fmt = f"{fecha_str[:4]}-{fecha_str[4:6]}-{fecha_str[6:]}" if len(fecha_str)==8 else fecha_str
    conn      = get_db()
    conn.execute("UPDATE indicadores SET ibc=0 WHERE fecha=?", (fecha_fmt,))
    conn.commit(); conn.close()
    return jsonify({"status": "ok"})

@app.route('/admin/historial_indices')
@login_required
def admin_historial_indices():
    """Últimos registros de la tabla indices para mostrar en panel admin"""
    limit = request.args.get('limit', 15, type=int)
    conn  = get_db()
    rows  = conn.execute(
        "SELECT fecha, ibc as valor FROM indicadores WHERE ibc > 0 ORDER BY fecha DESC LIMIT ?",
        (limit,)
    ).fetchall()
    conn.close()
    return jsonify([{
        'fecha':     r['fecha'],
        'valor':     r['valor'],
        'variacion': 0,
        'fuente':    'manual'
    } for r in rows])



# ========== API: HISTÓRICO DE ACCIONES (para consulta.html) ==========

@app.route('/api/historico')
@api_permiso_requerido('ver_consulta')
def api_historico():
    """
    Devuelve el histórico completo de una acción desde la tabla 'mercado'.
    Responde con { datos: [...], actual: {...} }
    Cada registro: { fecha, simbolo, nombre, apertura, maximo, minimo, cierre, volumen, variacion }
    """
    simbolo = request.args.get('simbolo', '').upper().strip()
    if not simbolo:
        return jsonify({"error": "Parámetro 'simbolo' requerido"}), 400

    conn = get_db()
    rows = conn.execute("""
        SELECT fecha, simbolo, nombre, apertura, maximo, minimo, cierre, volumen, variacion
        FROM mercado
        WHERE simbolo = ? AND cierre > 0
        ORDER BY fecha ASC
    """, (simbolo,)).fetchall()
    conn.close()

    datos = [dict(r) for r in rows]
    actual = datos[-1] if datos else {}

    return jsonify({"datos": datos, "actual": actual})


@app.route('/api/libro-ordenes/<simbolo>')
@api_permiso_requerido('ver_consulta')
def api_libro_ordenes(simbolo):
    """
    Proxy al libro de órdenes de la BVC para un símbolo dado.
    Responde con { success: bool, datos: [...] }
    """
    simbolo = simbolo.upper().strip()
    datos_bvc = obtener_datos_bvc(simbolo)
    if datos_bvc is None:
        return jsonify({"success": False, "datos": [], "error": "No se pudo conectar con la BVC"}), 200

    # La API BVC devuelve una lista directamente o un dict con 'data'/'ordenes'
    if isinstance(datos_bvc, list):
        return jsonify({"success": True, "datos": datos_bvc})
    elif isinstance(datos_bvc, dict):
        # Intentar extraer la lista de órdenes del formato conocido
        for key in ('ordenes', 'data', 'libroOrdenes', 'items'):
            if key in datos_bvc and isinstance(datos_bvc[key], list):
                return jsonify({"success": True, "datos": datos_bvc[key]})
        # Si no encontramos la clave, devolver el dict como lista de un elemento
        return jsonify({"success": True, "datos": [datos_bvc]})

    return jsonify({"success": False, "datos": [], "error": "Formato de respuesta inesperado"}), 200


# ========== ADMIN: DÓLAR BCV (tabla indicadores) ==========

@app.route('/admin/guardar_dolar', methods=['POST'])
@login_required
def admin_guardar_dolar():
    """Guarda o actualiza el dólar BCV en la columna 'dolar' de tabla indicadores"""
    data  = request.json
    fecha = data.get('fecha', '')   # viene como YYYY-MM-DD
    valor = float(data.get('valor', 0))
    if not fecha or not valor:
        return jsonify({"status": "error", "message": "Fecha y valor son requeridos"}), 400
    conn = get_db()
    try:
        conn.execute('''
            INSERT INTO indicadores (fecha, dolar, ibc)
            VALUES (?, ?, 0)
            ON CONFLICT(fecha) DO UPDATE SET dolar = excluded.dolar
        ''', (fecha, valor))
        conn.commit()
        conn.close()
        return jsonify({"status": "ok", "fecha": fecha, "valor": valor})
    except Exception as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/admin/obtener_historial_indices/<tipo>')
@login_required
def admin_obtener_historial_indices(tipo):
    """Devuelve los últimos registros de dólar o ibc de la tabla indicadores"""
    conn = get_db()
    if tipo == 'dolar':
        rows = conn.execute(
            "SELECT fecha, dolar as valor FROM indicadores WHERE dolar > 0 ORDER BY fecha DESC LIMIT 15"
        ).fetchall()
    elif tipo == 'ibc':
        rows = conn.execute(
            "SELECT fecha, ibc as valor FROM indicadores WHERE ibc > 0 ORDER BY fecha DESC LIMIT 15"
        ).fetchall()
    else:
        conn.close()
        return jsonify([])
    conn.close()
    return jsonify([{"fecha": r["fecha"], "valor": r["valor"]} for r in rows])


@app.route('/admin/eliminar_indice', methods=['POST'])
@login_required
def admin_eliminar_indice():
    """Pone en 0 el dólar o el ibc de un registro en tabla indicadores"""
    data  = request.json
    tipo  = data.get('tipo')   # 'dolar' o 'ibc'
    fecha = data.get('fecha')  # YYYY-MM-DD
    if tipo not in ('dolar', 'ibc') or not fecha:
        return jsonify({"status": "error", "message": "Parámetros inválidos"}), 400
    conn = get_db()
    try:
        conn.execute(f"UPDATE indicadores SET {tipo} = 0 WHERE fecha = ?", (fecha,))
        conn.commit()
        conn.close()
        return jsonify({"status": "ok"})
    except Exception as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)}), 500



# ========== ADMIN: ENDPOINTS FALTANTES ==========

@app.route('/admin/obtener_registros/<fecha>')
@login_required
def admin_obtener_registros(fecha):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM mercado WHERE fecha = ? ORDER BY simbolo", (fecha,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/admin/guardar_masivo', methods=['POST'])
@login_required
def admin_guardar_masivo():
    data = request.json
    if not data or not isinstance(data, list):
        return jsonify({"status": "error", "message": "Datos inválidos"}), 400
    conn = get_db()
    try:
        for reg in data:
            fecha    = reg.get('fecha', '')
            simbolo  = reg.get('simbolo', '').upper().strip()
            apertura = float(reg.get('apertura', 0) or 0)
            maximo   = float(reg.get('maximo', 0) or 0)
            minimo   = float(reg.get('minimo', 0) or 0)
            cierre   = float(reg.get('cierre', 0) or 0)
            volumen  = float(reg.get('volumen', 0) or 0)
            variacion = ((cierre - apertura) / apertura) * 100 if apertura > 0 else 0
            if not fecha or not simbolo:
                continue
            conn.execute("""
                INSERT OR REPLACE INTO mercado
                (fecha, simbolo, nombre, apertura, maximo, minimo, cierre, volumen, variacion)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (fecha, simbolo, reg.get('nombre', simbolo), apertura, maximo, minimo, cierre, volumen, variacion))
        conn.commit()
        conn.close()
        return jsonify({"status": "ok", "guardados": len(data)})
    except Exception as e:
        conn.close()
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/admin/eliminar_registro', methods=['POST'])
@login_required
def admin_eliminar_registro():
    data    = request.json
    fecha   = data.get('fecha', '')
    simbolo = data.get('simbolo', '').upper().strip()
    if not fecha or not simbolo:
        return jsonify({"status": "error", "message": "Fecha y símbolo requeridos"}), 400
    conn = get_db()
    conn.execute("DELETE FROM mercado WHERE fecha = ? AND simbolo = ?", (fecha, simbolo))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})


@app.route('/admin/historial_dolar')
@login_required
def admin_historial_dolar():
    limit = request.args.get('limit', 10, type=int)
    conn  = get_db()
    rows  = conn.execute(
        "SELECT fecha, dolar as valor FROM indicadores WHERE dolar > 0 ORDER BY fecha DESC LIMIT ?",
        (limit,)
    ).fetchall()
    conn.close()
    return jsonify([{"fecha": r["fecha"], "valor": r["valor"]} for r in rows])


@app.route('/admin/cargar_masivo', methods=['GET', 'POST'])
@login_required
def admin_cargar_masivo():
    if request.method == 'GET':
        return redirect(url_for('admin_panel'))
    archivos = request.files.getlist('archivos')
    procesados = 0
    errores = []
    for archivo in archivos:
        if archivo and archivo.filename.lower().endswith('.dat'):
            ruta_tmp = os.path.join(DATA_DIR, archivo.filename)
            os.makedirs(DATA_DIR, exist_ok=True)
            archivo.save(ruta_tmp)
            if data_manager.procesar_dat(ruta_tmp):
                procesados += 1
            else:
                errores.append(archivo.filename)
    return jsonify({"status": "ok", "procesados": procesados, "errores": errores})

@app.route('/perfil', methods=['GET', 'POST'])
def perfil():
    if 'user_logged_in' not in session and 'admin_logged_in' not in session:
        return redirect(url_for('login'))
    # Admin tiene su propio panel para cambiar clave
    if 'admin_logged_in' in session:
        return redirect(url_for('admin_panel'))

    mensaje_ok = None
    mensaje_err = None

    if request.method == 'POST':
        clave_actual = request.form.get('clave_actual', '')
        clave_nueva  = request.form.get('clave_nueva', '')
        clave_conf   = request.form.get('clave_confirmacion', '')
        username = session.get('username')

        conn = get_db()
        user = conn.execute("SELECT * FROM usuarios WHERE username=?", (username,)).fetchone()
        conn.close()

        if not user or user['password'] != hash_password(clave_actual):
            mensaje_err = 'La contraseña actual es incorrecta.'
        elif len(clave_nueva) < 6:
            mensaje_err = 'La nueva contraseña debe tener al menos 6 caracteres.'
        elif clave_nueva != clave_conf:
            mensaje_err = 'Las contraseñas nuevas no coinciden.'
        else:
            conn = get_db()
            conn.execute("UPDATE usuarios SET password=? WHERE username=?",
                         (hash_password(clave_nueva), username))
            conn.commit(); conn.close()
            mensaje_ok = '✅ Contraseña actualizada correctamente.'

    return render_template('perfil.html', mensaje_ok=mensaje_ok, mensaje_err=mensaje_err)


# ========== PORTAFOLIO DE INVERSIONES ==========

INV_DB_PATH = os.path.join(BASE_DIR, "database", "inversiones.db")

def get_inv_db():
    conn = sqlite3.connect(INV_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def crear_tablas_inversiones():
    os.makedirs(os.path.dirname(INV_DB_PATH), exist_ok=True)
    conn = sqlite3.connect(INV_DB_PATH)
    conn.execute('''CREATE TABLE IF NOT EXISTS accion (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner TEXT NOT NULL DEFAULT '',
        fecha_compra TEXT NOT NULL,
        simbolo TEXT NOT NULL,
        nombre_accion TEXT,
        cantidad INTEGER NOT NULL,
        precio_compra REAL NOT NULL,
        monto_bruto REAL NOT NULL,
        derecho_registro REAL DEFAULT 0,
        comision REAL DEFAULT 0,
        iva REAL DEFAULT 0,
        monto_total REAL NOT NULL,
        fecha_registro TEXT DEFAULT CURRENT_TIMESTAMP,
        imagen_url TEXT DEFAULT ''
    )''')
    try:
        conn.execute("ALTER TABLE accion ADD COLUMN owner TEXT NOT NULL DEFAULT ''")
        conn.commit()
    except Exception:
        pass

    conn.execute('''CREATE TABLE IF NOT EXISTS venta (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        owner TEXT NOT NULL DEFAULT '',
        fecha_venta TEXT NOT NULL,
        simbolo TEXT NOT NULL,
        nombre_accion TEXT,
        cantidad INTEGER NOT NULL,
        precio_venta REAL NOT NULL,
        monto_bruto REAL NOT NULL,
        comision REAL DEFAULT 0,
        iva REAL DEFAULT 0,
        monto_neto REAL NOT NULL,
        costo_promedio REAL DEFAULT 0,
        ganancia_realizada REAL DEFAULT 0,
        fecha_registro TEXT DEFAULT CURRENT_TIMESTAMP
    )''')

    conn.execute('''CREATE TABLE IF NOT EXISTS precio_actual (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        simbolo TEXT UNIQUE NOT NULL,
        precio_actual REAL DEFAULT 0,
        variacion REAL DEFAULT 0,
        fecha_actualizacion TEXT DEFAULT CURRENT_TIMESTAMP,
        imagen_url TEXT DEFAULT ''
    )''')
    conn.commit()
    conn.close()

crear_tablas_inversiones()

IMAGENES_ACCIONES = {
    'BVCC': 'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/banco-de-venezuela.png',
    'BVC':  'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/banco-de-venezuela.png',
    'BDV':  'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/banco-de-venezuela.png',
    'BOD':  'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/bod.png',
    'BFC':  'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/banfoandes.png',
    'BNC':  'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/bnc.png',
    'FNC':  'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/fondo-nacional-del-cafe.png',
    'TEL':  'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/cantv.png',
    'ECA':  'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/electricidad-de-caracas.png',
    'EDC':  'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/enelven.png',
    'MHO':  'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/mavesa.png',
    'RST':  'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/risto.png',
    'PDV':  'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/pdvsa.png',
    'CVG':  'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/cvg.png',
}
IMG_DEFAULT = 'https://www.bolsadecaracas.com/wp-content/uploads/2020/06/bolsa-de-caracas.png'

def obtener_imagen_accion(simbolo):
    """Devuelve la URL estática local si existe el PNG, si no None (el template usará las iniciales)."""
    if os.path.exists(os.path.join(BASE_DIR, f"static/img/acciones/{simbolo.upper()}.png")):
        return f"/static/img/acciones/{simbolo.upper()}.png"
    return None

app.jinja_env.globals['obtener_imagen_accion'] = obtener_imagen_accion

def calcular_resumen_inversiones(owner=None):
    conn_inv = get_inv_db()
    q_owner = "WHERE owner=?" if owner else ""
    p_owner = (owner,) if owner else ()

    compras = [dict(r) for r in conn_inv.execute(
        f"SELECT * FROM accion {q_owner}", p_owner).fetchall()]
    ventas  = [dict(r) for r in conn_inv.execute(
        f"SELECT * FROM venta {q_owner}", p_owner).fetchall()]
    conn_inv.close()

    # Últimos precios de mercado
    conn_bvc = get_db()
    precios_mercado = {}
    for r in conn_bvc.execute("""
        SELECT m.simbolo, m.cierre, m.variacion, m.nombre, m.fecha
        FROM mercado m
        INNER JOIN (SELECT simbolo, MAX(fecha) as max_fecha FROM mercado GROUP BY simbolo) lat
            ON m.simbolo=lat.simbolo AND m.fecha=lat.max_fecha
    """).fetchall():
        precios_mercado[r['simbolo']] = {
            'precio': r['cierre'], 'variacion': r['variacion'],
            'nombre': r['nombre'], 'fecha': r['fecha']
        }
    conn_bvc.close()

    # Agrupar compras por símbolo
    compras_sym = {}
    for c in compras:
        s = c['simbolo']
        if s not in compras_sym:
            compras_sym[s] = {'cant': 0, 'costo': 0.0, 'nombre': c['nombre_accion']}
        compras_sym[s]['cant']  += c['cantidad']
        compras_sym[s]['costo'] += c['monto_total']

    # Agrupar ventas por símbolo (cantidad vendida y ganancia realizada)
    ventas_sym = {}
    ganancia_realizada_total = 0.0
    for v in ventas:
        s = v['simbolo']
        if s not in ventas_sym:
            ventas_sym[s] = {'cant': 0, 'ganancia': 0.0}
        ventas_sym[s]['cant']    += v['cantidad']
        ventas_sym[s]['ganancia'] += v.get('ganancia_realizada', 0.0)
        ganancia_realizada_total  += v.get('ganancia_realizada', 0.0)

    # Construir posiciones netas (compras − ventas)
    resumen = {
        'total_invertido': 0.0, 'valor_actual': 0.0,
        'ganancia_perdida': 0.0, 'rendimiento_porcentaje': 0.0,
        'ganancia_realizada': ganancia_realizada_total,
        'detalle_acciones': [], 'cantidad_acciones': 0, 'acciones_diferentes': 0
    }

    todos_simbolos = set(compras_sym.keys()) | set(ventas_sym.keys())
    for sym in todos_simbolos:
        c_data = compras_sym.get(sym, {'cant': 0, 'costo': 0.0, 'nombre': sym})
        v_data = ventas_sym.get(sym, {'cant': 0, 'ganancia': 0.0})

        cant_neta = c_data['cant'] - v_data['cant']
        if cant_neta <= 0:
            continue  # posición completamente cerrada → no aparece en posiciones abiertas

        # Costo proporcional de la posición restante (FIFO simplificado por promedio)
        costo_promedio = c_data['costo'] / c_data['cant'] if c_data['cant'] > 0 else 0
        costo_neto = costo_promedio * cant_neta

        pm = precios_mercado.get(sym, {})
        precio_act = pm.get('precio', 0.0)
        variacion_diaria = pm.get('variacion', 0.0)
        fecha_precio = pm.get('fecha', '—')
        nombre_mercado = pm.get('nombre') or c_data['nombre'] or sym

        v_actual = cant_neta * precio_act
        ganancia_no_realizada = v_actual - costo_neto
        rendimiento = (ganancia_no_realizada / costo_neto * 100) if costo_neto > 0 else 0

        resumen['total_invertido']   += costo_neto
        resumen['valor_actual']      += v_actual
        resumen['cantidad_acciones'] += cant_neta

        resumen['detalle_acciones'].append({
            'simbolo': sym, 'nombre': nombre_mercado,
            'cantidad': cant_neta,
            'costo_total': costo_neto,
            'precio_promedio_por_accion': costo_promedio,
            'precio_actual': precio_act,
            'fecha_precio': fecha_precio,
            'valor_actual': v_actual,
            'ganancia': ganancia_no_realizada,
            'rendimiento': rendimiento,
            'variacion_diaria': variacion_diaria,
            'imagen_url': obtener_imagen_accion(sym),
            'tiene_imagen': obtener_imagen_accion(sym) is not None,
            'cant_vendida': v_data['cant'],
        })

    resumen['acciones_diferentes']  = len(resumen['detalle_acciones'])
    resumen['ganancia_perdida']      = resumen['valor_actual'] - resumen['total_invertido']
    if resumen['total_invertido'] > 0:
        resumen['rendimiento_porcentaje'] = (resumen['ganancia_perdida'] / resumen['total_invertido']) * 100
    return resumen

# --- Rutas portafolio ---

@app.route('/portafolio')
@permiso_requerido('ver_portafolio')
def portafolio():
    owner = session.get('username', '')
    resumen = calcular_resumen_inversiones(owner=owner)
    conn = get_inv_db()
    transacciones = [dict(r) for r in conn.execute(
        "SELECT * FROM accion WHERE owner=? ORDER BY fecha_registro DESC", (owner,)
    ).fetchall()]
    ventas = [dict(r) for r in conn.execute(
        "SELECT * FROM venta WHERE owner=? ORDER BY fecha_registro DESC", (owner,)
    ).fetchall()]
    conn.close()
    return render_template('portafolio.html', resumen=resumen,
                           transacciones=transacciones, ventas=ventas)


@app.route('/portafolio/vender', methods=['GET', 'POST'])
@permiso_requerido('ver_portafolio')
def vender_accion():
    owner = session.get('username', '')

    if request.method == 'POST':
        try:
            simbolo  = request.form['simbolo'].strip().upper()
            cantidad = int(request.form['cantidad'])
            precio_venta = float(request.form['precio_venta'])
            fecha_venta  = request.form['fecha_venta']
            comision = float(request.form.get('comision', 0) or 0)
            iva      = float(request.form.get('iva', 0) or 0)

            # Verificar que el usuario tiene suficientes acciones
            conn = get_inv_db()
            compras_rows = conn.execute(
                "SELECT SUM(cantidad) as total, SUM(monto_total) as costo FROM accion WHERE owner=? AND simbolo=?",
                (owner, simbolo)
            ).fetchone()
            ventas_rows = conn.execute(
                "SELECT SUM(cantidad) as total FROM venta WHERE owner=? AND simbolo=?",
                (owner, simbolo)
            ).fetchone()

            cant_comprada = compras_rows['total'] or 0
            costo_total   = compras_rows['costo'] or 0.0
            cant_vendida  = ventas_rows['total'] or 0
            cant_disponible = cant_comprada - cant_vendida

            if cantidad > cant_disponible:
                conn.close()
                from flask import flash
                flash(f"No puedes vender {cantidad} acciones. Solo tienes {cant_disponible} disponibles de {simbolo}.", "danger")
                return redirect(url_for('vender_accion'))

            costo_promedio = costo_total / cant_comprada if cant_comprada > 0 else 0
            monto_bruto    = cantidad * precio_venta
            monto_neto     = monto_bruto - comision - iva
            ganancia       = monto_neto - (costo_promedio * cantidad)

            nombre_accion = conn.execute(
                "SELECT nombre_accion FROM accion WHERE owner=? AND simbolo=? LIMIT 1",
                (owner, simbolo)
            ).fetchone()
            nombre = nombre_accion['nombre_accion'] if nombre_accion else simbolo

            conn.execute('''INSERT INTO venta
                (owner, fecha_venta, simbolo, nombre_accion, cantidad, precio_venta,
                 monto_bruto, comision, iva, monto_neto, costo_promedio, ganancia_realizada)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''',
                (owner, fecha_venta, simbolo, nombre, cantidad, precio_venta,
                 monto_bruto, comision, iva, monto_neto, costo_promedio, ganancia))
            conn.commit()
            conn.close()

            from flask import flash
            signo = '+' if ganancia >= 0 else ''
            flash(f"Venta de {cantidad} {simbolo} registrada. Ganancia realizada: {signo}Bs. {ganancia:,.2f}", "success")
            return redirect(url_for('portafolio'))

        except Exception as e:
            from flask import flash
            flash(f"Error al registrar venta: {e}", "danger")

    # GET — construir lista de posiciones disponibles para vender
    conn = get_inv_db()
    rows_c = conn.execute(
        "SELECT simbolo, nombre_accion, SUM(cantidad) as cant, SUM(monto_total) as costo FROM accion WHERE owner=? GROUP BY simbolo",
        (owner,)
    ).fetchall()
    rows_v = conn.execute(
        "SELECT simbolo, SUM(cantidad) as cant FROM venta WHERE owner=? GROUP BY simbolo",
        (owner,)
    ).fetchall()
    conn.close()

    vendidas = {r['simbolo']: r['cant'] for r in rows_v}
    posiciones = []
    for r in rows_c:
        disponible = r['cant'] - vendidas.get(r['simbolo'], 0)
        if disponible > 0:
            # último precio del mercado
            conn_bvc = get_db()
            pm = conn_bvc.execute(
                "SELECT cierre FROM mercado WHERE simbolo=? ORDER BY fecha DESC LIMIT 1",
                (r['simbolo'],)
            ).fetchone()
            conn_bvc.close()
            posiciones.append({
                'simbolo': r['simbolo'],
                'nombre': r['nombre_accion'] or r['simbolo'],
                'disponible': disponible,
                'costo_promedio': round(r['costo'] / r['cant'], 4) if r['cant'] > 0 else 0,
                'precio_mercado': pm['cierre'] if pm else 0,
                'imagen_url': obtener_imagen_accion(r['simbolo']),
            })

    return render_template('portafolio/vender_accion.html', posiciones=posiciones)


@app.route('/portafolio/venta/eliminar/<int:id>', methods=['POST'])
@permiso_requerido('ver_portafolio')
def eliminar_venta(id):
    owner = session.get('username', '')
    conn = get_inv_db()
    conn.execute("DELETE FROM venta WHERE id=? AND owner=?", (id, owner))
    conn.commit(); conn.close()
    from flask import flash
    flash("Venta eliminada correctamente.", "success")
    return redirect(url_for('portafolio'))

@app.route('/portafolio/agregar', methods=['GET', 'POST'])
@permiso_requerido('ver_portafolio')
def agregar_accion():
    if request.method == 'POST':
        try:
            accion_info = request.form.get('accion_seleccionada', '')
            simbolo, nombre = (accion_info.split(' - ', 1) + [''])[:2]
            simbolo = simbolo.strip()
            nombre = nombre.strip()
            conn = get_inv_db()
            conn.execute('''INSERT INTO accion
                (owner, fecha_compra, simbolo, nombre_accion, cantidad, precio_compra,
                 monto_bruto, derecho_registro, comision, iva, monto_total, imagen_url)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)''', (
                session.get('username', ''),
                request.form['fecha_compra'], simbolo, nombre,
                int(request.form['cantidad']),
                float(request.form['precio_compra']),
                float(request.form['monto_bruto']),
                float(request.form.get('derecho_registro', 0) or 0),
                float(request.form.get('comision', 0) or 0),
                float(request.form.get('iva', 0) or 0),
                float(request.form['monto_total']),
                obtener_imagen_accion(simbolo)
            ))
            conn.commit()
            conn.close()
            from flask import flash
            flash("Inversión agregada correctamente", "success")
            return redirect(url_for('portafolio'))
        except Exception as e:
            from flask import flash
            flash(f"Error al guardar: {e}", "danger")

    # Cargar acciones disponibles desde la BD local
    conn = get_db()
    ultima_fecha = conn.execute("SELECT MAX(fecha) FROM mercado").fetchone()[0]
    acciones_raw = []
    if ultima_fecha:
        rows = conn.execute(
            "SELECT simbolo, nombre, cierre, variacion, apertura, volumen FROM mercado WHERE fecha=? AND cierre>0 ORDER BY simbolo",
            (ultima_fecha,)
        ).fetchall()
        acciones_raw = [{'s': r['simbolo'], 'n': r['nombre'] or r['simbolo'],
                         'precio': r['cierre'], 'variacion': r['variacion'],
                         'apertura': r['apertura'], 'cantidad': r['volumen'], 'monto': 0} for r in rows]
    conn.close()
    return render_template('portafolio/agregar_accion.html', acciones_disponibles=acciones_raw)

@app.route('/portafolio/editar/<int:id>', methods=['GET', 'POST'])
@permiso_requerido('ver_portafolio')
def editar_accion(id):
    owner = session.get('username', '')
    conn = get_inv_db()
    accion = conn.execute("SELECT * FROM accion WHERE id=? AND owner=?", (id, owner)).fetchone()
    conn.close()
    if not accion:
        return redirect(url_for('portafolio'))
    accion = dict(accion)

    if request.method == 'POST':
        try:
            accion_info = request.form.get('accion_seleccionada', '')
            simbolo, nombre = (accion_info.split(' - ', 1) + [''])[:2]
            simbolo = simbolo.strip(); nombre = nombre.strip()
            conn = get_inv_db()
            conn.execute('''UPDATE accion SET
                fecha_compra=?, simbolo=?, nombre_accion=?, cantidad=?, precio_compra=?,
                monto_bruto=?, derecho_registro=?, comision=?, iva=?, monto_total=?, imagen_url=?
                WHERE id=?''', (
                request.form['fecha_compra'], simbolo, nombre,
                int(request.form['cantidad']),
                float(request.form['precio_compra']),
                float(request.form['monto_bruto']),
                float(request.form.get('derecho_registro', 0) or 0),
                float(request.form.get('comision', 0) or 0),
                float(request.form.get('iva', 0) or 0),
                float(request.form['monto_total']),
                obtener_imagen_accion(simbolo), id
            ))
            conn.commit(); conn.close()
            from flask import flash
            flash("Inversión actualizada correctamente", "success")
            return redirect(url_for('portafolio'))
        except Exception as e:
            from flask import flash
            flash(f"Error al actualizar: {e}", "danger")

    conn = get_db()
    ultima_fecha = conn.execute("SELECT MAX(fecha) FROM mercado").fetchone()[0]
    acciones_raw = []
    if ultima_fecha:
        rows = conn.execute(
            "SELECT simbolo, nombre, cierre, variacion, apertura, volumen FROM mercado WHERE fecha=? AND cierre>0 ORDER BY simbolo",
            (ultima_fecha,)
        ).fetchall()
        acciones_raw = [{'s': r['simbolo'], 'n': r['nombre'] or r['simbolo'],
                         'precio': r['cierre'], 'variacion': r['variacion'],
                         'apertura': r['apertura'], 'cantidad': r['volumen'], 'monto': 0} for r in rows]
    conn.close()
    return render_template('portafolio/editar_accion.html', accion=accion, acciones_disponibles=acciones_raw)

@app.route('/portafolio/eliminar/<int:id>', methods=['POST'])
@permiso_requerido('ver_portafolio')
def eliminar_accion(id):
    owner = session.get('username', '')
    conn = get_inv_db()
    conn.execute("DELETE FROM accion WHERE id=? AND owner=?", (id, owner))
    conn.commit(); conn.close()
    from flask import flash
    flash("Inversión eliminada correctamente", "success")
    return redirect(url_for('portafolio'))

@app.route('/portafolio/actualizar-precios')
@login_required
def actualizar_precios_portafolio():
    """Sincroniza precios del portafolio desde la tabla mercado local"""
    conn_bvc = get_db()
    ultima_fecha = conn_bvc.execute("SELECT MAX(fecha) FROM mercado").fetchone()[0]
    if ultima_fecha:
        rows = conn_bvc.execute(
            "SELECT simbolo, cierre, variacion FROM mercado WHERE fecha=? AND cierre>0",
            (ultima_fecha,)
        ).fetchall()
        conn_bvc.close()
        conn_inv = get_inv_db()
        for r in rows:
            existing = conn_inv.execute(
                "SELECT id FROM precio_actual WHERE simbolo=?", (r['simbolo'],)
            ).fetchone()
            if existing:
                conn_inv.execute(
                    "UPDATE precio_actual SET precio_actual=?, variacion=?, fecha_actualizacion=datetime('now') WHERE simbolo=?",
                    (r['cierre'], r['variacion'], r['simbolo'])
                )
            else:
                conn_inv.execute(
                    "INSERT INTO precio_actual (simbolo, precio_actual, variacion, imagen_url) VALUES (?,?,?,?)",
                    (r['simbolo'], r['cierre'], r['variacion'], obtener_imagen_accion(r['simbolo']))
                )
        conn_inv.commit(); conn_inv.close()
        from flask import flash
        flash(f"Precios actualizados desde sesión del {ultima_fecha}", "success")
    else:
        conn_bvc.close()
        from flask import flash
        flash("No hay datos de mercado disponibles", "warning")
    return redirect(url_for('portafolio'))

@app.route('/portafolio/resumen')
@permiso_requerido('ver_portafolio')
def resumen_portafolio():
    owner = session.get('username', '')
    resumen = calcular_resumen_inversiones(owner=owner)
    return render_template('portafolio/resumen.html', resumen=resumen)

@app.route('/portafolio/info-calculos')
def info_calculos():
    return render_template('portafolio/info_calculos.html')

# ========== ANÁLISIS AVANZADO ==========

@app.route('/analisis-avanzado')
@permiso_requerido('ver_prediccion')
def analisis_avanzado():
    conn = get_db()
    simbolos = [dict(r) for r in conn.execute("""
        SELECT simbolo, nombre FROM mercado
        WHERE cierre > 0
        GROUP BY simbolo
        HAVING MAX(fecha) >= date('now', '-6 months')
        ORDER BY simbolo
    """).fetchall()]
    conn.close()
    return render_template('analisis_avanzado.html', simbolos=simbolos)


@app.route('/api/correlacion')
@permiso_requerido('ver_prediccion')
def api_correlacion():
    """
    Calcula la matriz de correlación de retornos diarios entre los símbolos
    seleccionados. Devuelve: labels[], matrix[][] (valores entre -1 y +1).
    """
    simbolos_param = request.args.get('simbolos', '')
    periodo        = int(request.args.get('periodo', 90))  # días hacia atrás

    simbolos = [s.strip().upper() for s in simbolos_param.split(',') if s.strip()]
    if len(simbolos) < 2:
        return jsonify({"error": "Se necesitan al menos 2 símbolos"}), 400

    from datetime import date, timedelta
    fecha_min = (date.today() - timedelta(days=periodo)).isoformat()
    placeholders = ','.join('?' * len(simbolos))

    conn = get_db()
    rows = conn.execute(
        f"SELECT fecha, simbolo, cierre FROM mercado WHERE simbolo IN ({placeholders}) AND fecha >= ? ORDER BY fecha ASC",
        simbolos + [fecha_min]
    ).fetchall()
    conn.close()

    # Agrupar cierres por símbolo
    from collections import defaultdict
    series = defaultdict(dict)
    for r in rows:
        series[r['simbolo']][r['fecha']] = r['cierre']

    # Intersección de fechas donde todos los símbolos tienen datos
    fechas_comunes = None
    for sym in simbolos:
        f = set(series[sym].keys())
        fechas_comunes = f if fechas_comunes is None else fechas_comunes & f
    fechas_comunes = sorted(fechas_comunes) if fechas_comunes else []

    if len(fechas_comunes) < 5:
        return jsonify({"error": "Datos insuficientes para correlación"}), 400

    # Retornos diarios por símbolo
    def retornos(sym):
        precios = [series[sym][f] for f in fechas_comunes]
        return [(precios[i] - precios[i-1]) / precios[i-1] for i in range(1, len(precios)) if precios[i-1] > 0]

    ret = {sym: retornos(sym) for sym in simbolos}
    n   = min(len(v) for v in ret.values())

    def mean(lst):   return sum(lst) / len(lst)
    def stddev(lst, m): return (sum((x-m)**2 for x in lst) / len(lst)) ** 0.5

    # Calcular correlación de Pearson par a par
    matrix = []
    for s1 in simbolos:
        row = []
        for s2 in simbolos:
            a = ret[s1][:n]; b = ret[s2][:n]
            ma, mb = mean(a), mean(b)
            sa, sb = stddev(a, ma), stddev(b, mb)
            if sa == 0 or sb == 0:
                row.append(1.0 if s1 == s2 else 0.0)
            else:
                cov = sum((a[i]-ma)*(b[i]-mb) for i in range(n)) / n
                row.append(round(cov / (sa * sb), 4))
        matrix.append(row)

    return jsonify({"labels": simbolos, "matrix": matrix, "n_dias": len(fechas_comunes)})


@app.route('/api/prediccion')
@permiso_requerido('ver_prediccion')
def api_prediccion():
    """
    Pronóstico simple para un símbolo usando:
    - Regresión lineal sobre los últimos `periodo` días (tendencia)
    - Media móvil exponencial (EMA) para suavizado
    - Banda de confianza basada en la desviación estándar de residuos
    Devuelve: histórico[] + forecast[] con banda superior e inferior.
    """
    simbolo    = request.args.get('simbolo', '').upper()
    periodo    = int(request.args.get('periodo', 60))   # días históricos
    horizonte  = int(request.args.get('horizonte', 20)) # días a proyectar

    if not simbolo:
        return jsonify({"error": "Símbolo requerido"}), 400

    from datetime import date, timedelta
    fecha_min = (date.today() - timedelta(days=periodo)).isoformat()

    conn = get_db()
    rows = conn.execute(
        "SELECT fecha, cierre FROM mercado WHERE simbolo=? AND fecha>=? ORDER BY fecha ASC",
        (simbolo, fecha_min)
    ).fetchone() and conn.execute(
        "SELECT fecha, cierre FROM mercado WHERE simbolo=? AND fecha>=? ORDER BY fecha ASC",
        (simbolo, fecha_min)
    ).fetchall()
    ultimo = conn.execute(
        "SELECT cierre FROM mercado WHERE simbolo=? ORDER BY fecha DESC LIMIT 1", (simbolo,)
    ).fetchone()
    conn.close()

    if not rows or len(rows) < 5:
        return jsonify({"error": "Datos insuficientes"}), 400

    fechas  = [r['fecha'] for r in rows]
    precios = [r['cierre'] for r in rows]
    n       = len(precios)

    # Regresión lineal simple (mínimos cuadrados)
    x      = list(range(n))
    xm     = sum(x) / n
    ym     = sum(precios) / n
    num    = sum((x[i]-xm) * (precios[i]-ym) for i in range(n))
    den    = sum((x[i]-xm)**2 for i in range(n))
    slope  = num / den if den != 0 else 0
    interc = ym - slope * xm

    # Residuos y desviación estándar para banda de confianza
    residuos = [precios[i] - (interc + slope * i) for i in range(n)]
    mean_r   = sum(residuos) / n
    std_r    = (sum((r - mean_r)**2 for r in residuos) / n) ** 0.5

    # EMA sobre histórico (alpha = 2/(periodo+1))
    alpha  = 2 / (min(n, 14) + 1)
    ema    = [precios[0]]
    for p in precios[1:]:
        ema.append(alpha * p + (1 - alpha) * ema[-1])

    # Generar fechas futuras (días hábiles aproximados: saltar sábado y domingo)
    from datetime import datetime as dt
    ultima_fecha = dt.strptime(fechas[-1], '%Y-%m-%d').date()
    fechas_fut   = []
    d_cursor     = ultima_fecha
    while len(fechas_fut) < horizonte:
        d_cursor += timedelta(days=1)
        if d_cursor.weekday() < 5:   # 0=lun … 4=vie
            fechas_fut.append(d_cursor.isoformat())

    # Proyección: continuar la tendencia lineal
    forecast  = []
    banda_sup = []
    banda_inf = []
    factor_incertidumbre = 1.0
    for j, f in enumerate(fechas_fut):
        xi      = n + j
        pred    = interc + slope * xi
        factor_incertidumbre += 0.08   # banda crece con el horizonte
        sup     = pred + std_r * factor_incertidumbre
        inf     = pred - std_r * factor_incertidumbre
        forecast.append(round(max(pred, 0), 4))
        banda_sup.append(round(max(sup, 0), 4))
        banda_inf.append(round(max(inf, 0), 4))

    return jsonify({
        "simbolo":    simbolo,
        "fechas_hist":   fechas,
        "precios_hist":  precios,
        "ema_hist":      [round(e, 4) for e in ema],
        "fechas_fut":    fechas_fut,
        "forecast":      forecast,
        "banda_sup":     banda_sup,
        "banda_inf":     banda_inf,
        "slope":         round(slope, 6),
        "precio_actual": ultimo['cierre'] if ultimo else 0,
        "precio_objetivo": round(forecast[-1], 2) if forecast else 0,
    })


# ========== API: TARJETA DE POSICIÓN ==========

@app.route('/api/portafolio/tarjeta/<simbolo>')
def api_tarjeta_posicion(simbolo):
    """Devuelve datos para la tarjeta compartible de una posición del usuario."""
    if 'user_logged_in' not in session and 'admin_logged_in' not in session:
        return jsonify({"error": "no_session"}), 401

    owner = session.get('username', '')
    simbolo = simbolo.upper()

    conn_inv = get_inv_db()
    compras = conn_inv.execute(
        "SELECT SUM(cantidad) as cant, SUM(monto_total) as costo FROM accion WHERE owner=? AND simbolo=?",
        (owner, simbolo)
    ).fetchone()
    ventas = conn_inv.execute(
        "SELECT SUM(cantidad) as cant FROM venta WHERE owner=? AND simbolo=?",
        (owner, simbolo)
    ).fetchone()
    nombre_row = conn_inv.execute(
        "SELECT nombre_accion FROM accion WHERE owner=? AND simbolo=? LIMIT 1",
        (owner, simbolo)
    ).fetchone()
    conn_inv.close()

    cant_comprada = compras['cant'] or 0
    costo_total   = compras['costo'] or 0.0
    cant_vendida  = ventas['cant'] or 0
    cant_neta     = cant_comprada - cant_vendida

    if cant_neta <= 0:
        return jsonify({"error": "sin_posicion"}), 404

    costo_promedio = costo_total / cant_comprada if cant_comprada > 0 else 0

    # Histórico de precios de mercado (máximo 1 año)
    conn_bvc = get_db()
    rows = conn_bvc.execute(
        "SELECT fecha, cierre, apertura, maximo, minimo FROM mercado WHERE simbolo=? ORDER BY fecha ASC",
        (simbolo,)
    ).fetchall()
    ultimo = conn_bvc.execute(
        "SELECT cierre, variacion FROM mercado WHERE simbolo=? ORDER BY fecha DESC LIMIT 1",
        (simbolo,)
    ).fetchone()
    conn_bvc.close()

    precio_actual = ultimo['cierre'] if ultimo else 0
    variacion_dia = ultimo['variacion'] if ultimo else 0
    valor_actual  = cant_neta * precio_actual
    ganancia_no_r = valor_actual - (costo_promedio * cant_neta)
    rendimiento   = (ganancia_no_r / (costo_promedio * cant_neta) * 100) if costo_promedio > 0 else 0

    historico = [{'fecha': r['fecha'], 'cierre': r['cierre']} for r in rows]

    imagen_url = obtener_imagen_accion(simbolo)
    nombre     = nombre_row['nombre_accion'] if nombre_row else simbolo

    return jsonify({
        'simbolo':       simbolo,
        'nombre':        nombre,
        'imagen_url':    imagen_url,
        'precio_actual': precio_actual,
        'variacion_dia': variacion_dia,
        'costo_promedio':round(costo_promedio, 4),
        'cantidad':      cant_neta,
        'valor_actual':  round(valor_actual, 2),
        'ganancia':      round(ganancia_no_r, 2),
        'rendimiento':   round(rendimiento, 2),
        'historico':     historico,
    })


# ========== API: EVOLUCIÓN DEL PORTAFOLIO ==========

@app.route('/api/portafolio/evolucion')
def api_portafolio_evolucion():
    """
    Devuelve dos series por fecha:
      - valor_mercado: cant_neta(símbolo) × cierre(fecha)
      - capital_invertido: costo_promedio(símbolo) × cant_neta(símbolo)  [línea plana por posición]
    Más estadísticas reales de G/P basadas en datos de inversiones del usuario.
    """
    if 'user_logged_in' not in session and 'admin_logged_in' not in session:
        return jsonify({"error": "no_session"}), 401

    owner   = session.get('username', '')
    periodo = request.args.get('periodo', '1m')

    from datetime import date, timedelta
    from collections import defaultdict
    hoy = date.today()

    conn_inv = get_inv_db()

    # Compras: por símbolo → cantidad total y costo total
    compras_rows = conn_inv.execute(
        "SELECT simbolo, SUM(cantidad) as cant, SUM(monto_total) as costo FROM accion WHERE owner=? GROUP BY simbolo",
        (owner,)
    ).fetchall()

    # Ventas: por símbolo → cantidad total vendida
    ventas_rows = conn_inv.execute(
        "SELECT simbolo, SUM(cantidad) as cant FROM venta WHERE owner=? GROUP BY simbolo",
        (owner,)
    ).fetchall()

    # Compras ordenadas por fecha para reconstruir capital acumulado en el tiempo
    compras_detalle = conn_inv.execute(
        "SELECT fecha_compra, simbolo, cantidad, monto_total FROM accion WHERE owner=? ORDER BY fecha_compra ASC",
        (owner,)
    ).fetchall()

    primera_compra = conn_inv.execute(
        "SELECT MIN(fecha_compra) as primera FROM accion WHERE owner=?", (owner,)
    ).fetchone()
    conn_inv.close()

    if not compras_rows:
        return jsonify({"fechas": [], "valores_mercado": [], "capital_invertido": [],
                        "gp_total": 0, "gp_pct": 0, "capital_total": 0, "valor_actual_total": 0})

    vendidas = {r['simbolo']: r['cant'] for r in ventas_rows}

    # Posiciones netas y costo promedio por símbolo
    posiciones = {}   # sym → {cant_neta, costo_promedio}
    for r in compras_rows:
        neta = r['cant'] - vendidas.get(r['simbolo'], 0)
        if neta > 0:
            costo_prom = r['costo'] / r['cant'] if r['cant'] > 0 else 0
            posiciones[r['simbolo']] = {'cant': neta, 'costo_prom': costo_prom,
                                         'capital': round(costo_prom * neta, 2)}

    if not posiciones:
        return jsonify({"fechas": [], "valores_mercado": [], "capital_invertido": [],
                        "gp_total": 0, "gp_pct": 0, "capital_total": 0, "valor_actual_total": 0})

    # Capital total real (suma de costos × cant_neta)
    capital_total = sum(v['capital'] for v in posiciones.values())

    # Fecha de primera compra real
    fecha_primera = primera_compra['primera'] if primera_compra and primera_compra['primera'] else '2000-01-01'

    fecha_periodo = None
    if   periodo == '1s': fecha_periodo = (hoy - timedelta(weeks=1)).isoformat()
    elif periodo == '1m': fecha_periodo = (hoy - timedelta(days=30)).isoformat()
    elif periodo == '3m': fecha_periodo = (hoy - timedelta(days=90)).isoformat()
    elif periodo == '1y': fecha_periodo = (hoy - timedelta(days=365)).isoformat()

    fecha_min = max(fecha_primera, fecha_periodo) if fecha_periodo else fecha_primera

    simbolos     = list(posiciones.keys())
    placeholders = ','.join('?' * len(simbolos))

    conn_bvc = get_db()
    rows = conn_bvc.execute(
        f"SELECT fecha, simbolo, cierre FROM mercado WHERE simbolo IN ({placeholders}) AND fecha >= ? ORDER BY fecha ASC",
        simbolos + [fecha_min]
    ).fetchall()
    conn_bvc.close()

    # Agrupar precios por fecha
    por_fecha = defaultdict(dict)
    for r in rows:
        por_fecha[r['fecha']][r['simbolo']] = r['cierre']

    # ── Capital acumulado por fecha ──────────────────────────────────────
    # Reconstruimos cuánto capital estaba invertido en cada sesión de mercado,
    # acumulando las compras reales en orden cronológico.
    # Cada compra agrega: cantidad × precio_compra al capital acumulado.

    # Lista de eventos de compra ordenados por fecha
    eventos_compra = sorted(
        [{'fecha': dict(r)['fecha_compra'], 'monto': dict(r)['monto_total']} for r in compras_detalle],
        key=lambda x: x['fecha']
    )

    # Para cada fecha del gráfico, calculamos el capital acumulado
    # sumando todas las compras cuya fecha_compra <= fecha_sesion
    fechas_grafico = sorted(por_fecha.keys())

    def capital_en_fecha(fecha_sesion):
        return sum(e['monto'] for e in eventos_compra if e['fecha'] <= fecha_sesion)

    fechas           = []
    valores_mercado  = []
    capital_serie    = []

    for fecha in fechas_grafico:
        precios_dia = por_fecha[fecha]
        vm = sum(
            posiciones[sym]['cant'] * precios_dia[sym]
            for sym in simbolos if sym in precios_dia
        )
        if vm > 0:
            cap_en_este_dia = capital_en_fecha(fecha)
            fechas.append(fecha)
            valores_mercado.append(round(vm, 2))
            capital_serie.append(round(cap_en_este_dia, 2))

    # G/P real = último valor mercado − capital invertido
    valor_actual_total = valores_mercado[-1] if valores_mercado else 0
    gp_total = round(valor_actual_total - capital_total, 2)
    gp_pct   = round(gp_total / capital_total * 100, 2) if capital_total > 0 else 0

    return jsonify({
        "fechas":            fechas,
        "valores_mercado":   valores_mercado,
        "capital_invertido": capital_serie,
        "gp_total":          gp_total,
        "gp_pct":            gp_pct,
        "capital_total":     round(capital_total, 2),
        "valor_actual_total":round(valor_actual_total, 2),
    })

# ========== ADMIN: CAMBIO DE CREDENCIALES ==========

@app.route('/admin/cambiar-clave', methods=['POST'])
@login_required
def admin_cambiar_clave():
    data          = request.json
    clave_actual  = data.get('clave_actual', '')
    clave_nueva   = data.get('clave_nueva', '')
    clave_conf    = data.get('clave_confirmacion', '')

    _, admin_pass_hash = get_admin_credenciales()
    if hash_password(clave_actual) != admin_pass_hash:
        return jsonify({"status": "error", "message": "La clave actual es incorrecta"}), 400
    if len(clave_nueva) < 6:
        return jsonify({"status": "error", "message": "La nueva clave debe tener al menos 6 caracteres"}), 400
    if clave_nueva != clave_conf:
        return jsonify({"status": "error", "message": "Las claves nuevas no coinciden"}), 400

    conn = sqlite3.connect(DB_PATH)
    conn.execute("UPDATE admin_credenciales SET password=? WHERE id=1", (hash_password(clave_nueva),))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok", "message": "Clave actualizada correctamente"})


@app.route('/admin/cambiar-usuario', methods=['POST'])
@login_required
def admin_cambiar_usuario():
    data         = request.json
    clave_actual = data.get('clave_actual', '')
    nuevo_user   = data.get('nuevo_usuario', '').strip()

    _, admin_pass_hash = get_admin_credenciales()
    if hash_password(clave_actual) != admin_pass_hash:
        return jsonify({"status": "error", "message": "La clave actual es incorrecta"}), 400
    if not nuevo_user or len(nuevo_user) < 3:
        return jsonify({"status": "error", "message": "El usuario debe tener al menos 3 caracteres"}), 400

    conn = sqlite3.connect(DB_PATH)
    conn.execute("UPDATE admin_credenciales SET username=? WHERE id=1", (nuevo_user,))
    conn.commit()
    conn.close()
    session['username'] = nuevo_user
    return jsonify({"status": "ok", "message": f"Usuario cambiado a '{nuevo_user}'"})


@app.route('/admin/usuarios/reset-clave', methods=['POST'])
@login_required
def admin_reset_clave_usuario():
    import random, string
    data    = request.json
    user_id = data.get('user_id')
    nueva   = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
    conn    = get_db()
    conn.execute("UPDATE usuarios SET password=? WHERE id=?", (hash_password(nueva), user_id))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok", "nueva_clave": nueva})


@app.route('/admin/backup')
@login_required
def admin_backup():
    import shutil, io
    backup_path = os.path.join(BASE_DIR, "database", "backup_bolsa.db")
    shutil.copy2(DB_PATH, backup_path)
    with open(backup_path, 'rb') as f:
        data = f.read()
    from flask import Response
    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
    return Response(
        data,
        mimetype='application/octet-stream',
        headers={"Content-Disposition": f"attachment; filename=bolsa_backup_{ts}.db"}
    )



if __name__ == '__main__':
    os.makedirs(DATA_DIR, exist_ok=True)
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=True, host='0.0.0.0', port=port)